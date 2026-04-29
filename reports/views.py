from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from core.models import EngineeringTeam, Department
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl


@login_required
def reports_index(request):
    teams = EngineeringTeam.objects.all()
    departments = Department.objects.all()
    total_teams = teams.count()
    total_departments = departments.count()
    teams_without_managers = teams.filter(manager=None)

    context = {
        'teams': teams,
        'departments': departments,
        'total_teams': total_teams,
        'total_departments': total_departments,
        'teams_without_managers': teams_without_managers,
    }
    return render(request, 'reports/index.html', context)


@login_required
def department_summary(request):
    departments = Department.objects.all()
    summary = []
    for dept in departments:
        teams = EngineeringTeam.objects.filter(department=dept)
        summary.append({
            'department': dept,
            'team_count': teams.count(),
            'teams': teams,
        })

    context = {
        'summary': summary,
    }
    return render(request, 'reports/department_summary.html', context)


@login_required
def team_detail(request, team_id):
    team = get_object_or_404(EngineeringTeam, id=team_id)
    members = team.members.all()

    context = {
        'team': team,
        'members': members,
    }
    return render(request, 'reports/team_detail.html', context)


@login_required
def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'

    teams = EngineeringTeam.objects.all()
    departments = Department.objects.all()
    teams_without_managers = teams.filter(manager=None)

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 750, "Sky Engineering - Teams Report")

    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, 720, f"Total Teams: {teams.count()}")
    p.drawString(50, 700, f"Total Departments: {departments.count()}")
    p.drawString(50, 680, f"Teams Without Managers: {teams_without_managers.count()}")

    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, 650, "All Teams:")
    p.setFont("Helvetica", 11)
    y = 630
    for team in teams:
        p.drawString(70, y, f"- {team.team_name} | {team.department} | {team.status}")
        y -= 20
        if y < 100:
            p.showPage()
            y = 750

    p.showPage()
    p.save()
    return response


@login_required
def export_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

    teams = EngineeringTeam.objects.all()
    departments = Department.objects.all()
    teams_without_managers = teams.filter(manager=None)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Metric", "Value"])
    ws1.append(["Total Teams", teams.count()])
    ws1.append(["Total Departments", departments.count()])
    ws1.append(["Teams Without Managers", teams_without_managers.count()])

    ws2 = wb.create_sheet("All Teams")
    ws2.append(["Team Name", "Department", "Status"])
    for team in teams:
        ws2.append([team.team_name, str(team.department), team.status])

    ws3 = wb.create_sheet("Teams Without Managers")
    ws3.append(["Team Name", "Department"])
    for team in teams_without_managers:
        ws3.append([team.team_name, str(team.department)])

    wb.save(response)
    return response